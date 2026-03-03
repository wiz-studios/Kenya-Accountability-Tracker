#!/usr/bin/env python3
"""
Seed leaders table with official sources:
- National Assembly members: local workbook
  "List of Members by Parties 13th Parliament as at 02122025.xlsx"
- Senate county delegations: local workbook
  "Senators County Delegations as of 2nd December, 2025.xlsx"
- County governors: Council of Governors page
  https://cog.go.ke/current-governors/

Usage:
  python scripts/seed/seed_leaders_official.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[2]
MP_WORKBOOK = ROOT / "List of Members by Parties 13th Parliament as at 02122025.xlsx"
SENATE_WORKBOOK = ROOT / "Senators County Delegations as of 2nd December, 2025.xlsx"
COUNTY_DATA_TS = ROOT / "lib" / "complete-kenya-data.ts"

MANUAL_CONSTITUENCY_COUNTY = {
    "kamkunji": "Nairobi",
    "tiaty": "Baringo",
    "ijara": "Garissa",
    "jomvu": "Mombasa",
    "magarini": "Kilifi",
    "borabu": "Nyamira",
    "ndaragwa": "Nyandarua",
    "kangema": "Murang'a",
    "royasambu": "Nairobi",
    "roysambu": "Nairobi",
    "eldamaravine": "Baringo",
    "laisamis": "Marsabit",
    "bura": "Tana River",
    "voi": "Taita Taveta",
}

HONORIFIC_PREFIX_RE = re.compile(
    r"^(?:H\.?E\.?|Hon\.?|Dr\.?|Prof\.?|FCPA|Maj\.?\s*\(Rtd\)|Maj\.?|Rtd\.?)\s+",
    re.IGNORECASE,
)


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def normalize_spacing(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def convert_last_first_to_first_last(name: str) -> str:
    raw = normalize_spacing(name)
    if "," not in raw:
        return raw
    surname, rest = raw.split(",", 1)
    surname = normalize_spacing(surname)
    rest = normalize_spacing(rest)
    return normalize_spacing(f"{rest} {surname}")


def clean_governor_name(name: str) -> str:
    value = normalize_spacing(name)
    while True:
        updated = HONORIFIC_PREFIX_RE.sub("", value)
        if updated == value:
            break
        value = normalize_spacing(updated)
    # Trim honorific suffixes after comma (EGH, CBS, OGW...)
    value = re.sub(r",\s*[A-Z][A-Z ,.-]*$", "", value).strip()
    return value


def parse_county_mappings() -> Dict[str, str]:
    text = COUNTY_DATA_TS.read_text(encoding="utf-8")

    county_pairs = re.findall(
        r'id:\s*"([^"]+)",\s*\n\s*name:\s*"([^"]+)",\s*\n\s*code:\s*"',
        text,
    )
    county_by_id = {county_id: county_name for county_id, county_name in county_pairs}

    constituency_pairs = re.findall(
        r'name:\s*"([^"]+)",\s*\n\s*countyId:\s*"([^"]+)"',
        text,
    )

    result: Dict[str, str] = {}
    for constituency, county_id in constituency_pairs:
        county_name = county_by_id.get(county_id)
        if county_name:
            result[norm(constituency)] = county_name

    result.update(MANUAL_CONSTITUENCY_COUNTY)
    return result


def infer_county(constituency: str, constituency_map: Dict[str, str]) -> Tuple[str, str]:
    constituency_raw = normalize_spacing(constituency)

    cwr_match = re.match(r"^(.*?)\s*\(CWR\)\s*$", constituency_raw, re.IGNORECASE)
    if cwr_match:
        county = normalize_spacing(cwr_match.group(1))
        return county, constituency_raw

    if constituency_raw.lower() == "nominated":
        return "National", constituency_raw

    county = constituency_map.get(norm(constituency_raw))
    if county:
        return county, constituency_raw

    return "Unknown", constituency_raw


def parse_mp_records(constituency_map: Dict[str, str]) -> Tuple[List[dict], List[dict]]:
    wb = openpyxl.load_workbook(MP_WORKBOOK, data_only=True)
    records: List[dict] = []
    unknowns: List[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Table 13":
            continue

        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            number = ws.cell(row, 1).value
            name = ws.cell(row, 2).value
            constituency = ws.cell(row, 3).value
            party = ws.cell(row, 4).value

            if not isinstance(number, (int, float)):
                continue
            if not name or not constituency or not party:
                continue

            county, constituency_normalized = infer_county(str(constituency), constituency_map)
            full_name = convert_last_first_to_first_last(str(name))

            if county == "Unknown":
                unknowns.append(
                    {
                        "name": full_name,
                        "constituency": constituency_normalized,
                        "party": normalize_spacing(str(party)),
                    }
                )

            records.append(
                {
                    "name": full_name,
                    "position": "Member of Parliament",
                    "county": county,
                    "constituency": constituency_normalized,
                    "party": normalize_spacing(str(party)),
                    "term": "2022-2027",
                    "accountability_score": 70,
                }
            )

    return records, unknowns


def split_senator_names(raw: str) -> List[str]:
    value = (raw or "").replace("\r", "\n")
    value = value.replace("\u2013", "-").replace("\u2014", "-").replace("\u2015", "-").replace("\u2212", "-")
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?i)((?:delegation|member))\s+(?=[A-Z][A-Za-z'’().-]+,)", r"\1\n", value)

    parts = [item.strip() for item in re.split(r"[\n]+", value) if item and item.strip()]
    cleaned: List[str] = []

    for part in parts:
        part = re.sub(r"\s*-?\s*(Head\s*of\s*delegation|Member)\b", "", part, flags=re.IGNORECASE)
        part = re.sub(r"[-\u2013\u2014\u2015\u2212]+\s*$", "", part)
        part = normalize_spacing(part).strip(" -")
        if part:
            cleaned.append(part)

    return cleaned


def split_senator_parties(raw: str, expected_count: int) -> List[str]:
    lines = [item.strip() for item in re.split(r"[\r\n]+", raw or "") if item and item.strip()]
    tokens: List[str] = []
    for line in lines:
        tokens.extend([token for token in re.split(r"\s+", line) if token])

    parties = tokens or lines
    if len(parties) == 1 and expected_count > 1:
        parties = parties * expected_count
    if len(parties) < expected_count:
        fallback = parties[-1] if parties else "Unknown"
        parties.extend([fallback] * (expected_count - len(parties)))

    return parties[:expected_count]


def parse_senator_records() -> List[dict]:
    ws = openpyxl.load_workbook(SENATE_WORKBOOK, data_only=True)["Table 1"]
    records: List[dict] = []

    for row in range(5, ws.max_row + 1):
        number = ws.cell(row, 1).value
        name_blob = ws.cell(row, 2).value
        county = ws.cell(row, 3).value
        party_blob = ws.cell(row, 4).value

        if not number or not name_blob or not county:
            continue

        names = split_senator_names(str(name_blob))
        parties = split_senator_parties(str(party_blob or ""), len(names))
        county_name = normalize_spacing(str(county)).replace("Taita-Taveta", "Taita Taveta")

        for index, senator_name in enumerate(names):
            full_name = convert_last_first_to_first_last(senator_name)
            records.append(
                {
                    "name": full_name,
                    "position": "Senator",
                    "county": county_name,
                    "constituency": "County-wide",
                    "party": normalize_spacing(parties[index]) or "Unknown",
                    "term": "2022-2027",
                    "accountability_score": 72,
                }
            )

    return records


def parse_governor_records() -> List[dict]:
    response = requests.get("https://cog.go.ke/current-governors/", timeout=45)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    records: List[dict] = []
    for heading in soup.find_all("h3"):
        name = normalize_spacing(heading.get_text(" ", strip=True))
        parent_text = normalize_spacing(heading.parent.get_text(" ", strip=True))
        county_match = re.search(r"County:\s*([^\n\r]+)$", parent_text, re.IGNORECASE)
        if not county_match:
            continue

        county = normalize_spacing(county_match.group(1)).replace("Taita-Taveta", "Taita Taveta")
        cleaned_name = clean_governor_name(name)
        if not cleaned_name:
            continue

        records.append(
            {
                "name": cleaned_name,
                "position": "County Governor",
                "county": county,
                "constituency": "County-wide",
                "party": "Unknown",
                "term": "2022-2027",
                "accountability_score": 74,
            }
        )

    return records


def dedupe_records(records: List[dict]) -> List[dict]:
    seen = set()
    output = []
    for record in records:
        key = (
            norm(record["name"]),
            norm(record["position"]),
            norm(record["county"]),
            norm(record["constituency"]),
        )
        if key in seen:
            continue
        seen.add(key)
        output.append(record)
    return output


def chunked(rows: List[dict], size: int) -> List[List[dict]]:
    return [rows[index : index + size] for index in range(0, len(rows), size)]


def replace_leaders_table(records: List[dict], supabase_url: str, service_role_key: str) -> None:
    base_url = supabase_url.rstrip("/") + "/rest/v1/leaders"
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Content-Type": "application/json",
    }

    delete_response = requests.delete(
        f"{base_url}?id=not.is.null",
        headers={**headers, "Prefer": "return=minimal"},
        timeout=60,
    )
    if delete_response.status_code >= 300:
        raise RuntimeError(f"Failed to clear leaders table: {delete_response.status_code} {delete_response.text}")

    for index, batch in enumerate(chunked(records, 500), start=1):
        insert_response = requests.post(
            base_url,
            headers={**headers, "Prefer": "return=minimal"},
            data=json.dumps(batch),
            timeout=60,
        )
        if insert_response.status_code >= 300:
            raise RuntimeError(
                f"Failed to insert leaders batch {index}: {insert_response.status_code} {insert_response.text}"
            )


def main() -> None:
    load_env_file(ROOT / ".env.local")
    load_env_file(ROOT / ".env")

    supabase_url = (
        os.getenv("SUPABASE_URL")
        or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
        or os.getenv("TARGET_SUPABASE_URL")
        or ""
    ).strip()
    service_role_key = (
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        or os.getenv("TARGET_SUPABASE_SERVICE_ROLE_KEY")
        or ""
    ).strip()

    if not supabase_url or not service_role_key:
        print("Missing Supabase URL or service-role key in environment.", file=sys.stderr)
        sys.exit(1)

    if not MP_WORKBOOK.exists():
        print(f"Missing workbook: {MP_WORKBOOK}", file=sys.stderr)
        sys.exit(1)
    if not SENATE_WORKBOOK.exists():
        print(f"Missing workbook: {SENATE_WORKBOOK}", file=sys.stderr)
        sys.exit(1)

    constituency_map = parse_county_mappings()

    mp_records, unknown_constituencies = parse_mp_records(constituency_map)
    senator_records = parse_senator_records()
    governor_records = parse_governor_records()

    combined = dedupe_records([*mp_records, *senator_records, *governor_records])

    replace_leaders_table(combined, supabase_url, service_role_key)

    position_counts: Dict[str, int] = {}
    for record in combined:
        position_counts[record["position"]] = position_counts.get(record["position"], 0) + 1

    print("Leaders table seeded successfully.")
    print(f"Total records: {len(combined)}")
    for position, count in sorted(position_counts.items()):
        print(f"- {position}: {count}")

    if unknown_constituencies:
        print("")
        print(f"Unmapped MP constituencies: {len(unknown_constituencies)}")
        for item in unknown_constituencies[:25]:
            print(f"  - {item['name']} | {item['constituency']} | {item['party']}")


if __name__ == "__main__":
    main()
